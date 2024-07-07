import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# 定义主文件夹路径
main_folder = 'checkpoints'

# 定义存储结果的字典
data = {
    'Folder': [],
    'Accuracy': [],
    'Precision Ao Mango': [],
    'Recall Ao Mango': [],
    'F1 Score Ao Mango': [],
    'Precision GuiQi Mango': [],
    'Recall GuiQi Mango': [],
    'F1 Score GuiQi Mango': [],
    'Precision Jinhuang Mango': [],
    'Recall Jinhuang Mango': [],
    'F1 Score Jinhuang Mango': [],
    'Precision Tainong Mango': [],
    'Recall Tainong Mango': [],
    'F1 Score Tainong Mango': [],
    'Flops':[],
    'Params':[],

}

# 遍历主文件夹中的所有子文件夹
for root, dirs, files in os.walk(main_folder):
    for dir in dirs:
        # 构建test.txt文件的路径
        test_file_path = os.path.join(root, dir, 'test.txt')
        
        # 检查test.txt文件是否存在
        if os.path.exists(test_file_path):
            with open(test_file_path, 'r') as file:
                content = file.read()
                
                # 使用正则表达式提取数据
                accuracy_match = re.search(r'Accuracy:\s*([\d\.]+)', content)
                ao_mango_match = re.search(r'Class Ao Mango - Precision:\s*([\d\.]+), Recall:\s*([\d\.]+), F1 Score:\s*([\d\.]+)', content)
                guiqi_mango_match = re.search(r'Class GuiQi Mango - Precision:\s*([\d\.]+), Recall:\s*([\d\.]+), F1 Score:\s*([\d\.]+)', content)
                jinhuang_mango_match = re.search(r'Class Jinhuang Mango - Precision:\s*([\d\.]+), Recall:\s*([\d\.]+), F1 Score:\s*([\d\.]+)', content)
                tainong_mango_match = re.search(r'Class Tainong Mango - Precision:\s*([\d\.]+), Recall:\s*([\d\.]+), F1 Score:\s*([\d\.]+)', content)
                Flops_match = re.search(r'Flops:\s*(.*)', content)
                Params_match = re.search(r'Params:\s*(.*)', content)


                # 获取上上一级文件夹名称
                folder_name = os.path.basename(os.path.dirname(test_file_path))
                
                # 如果所有匹配项都找到了，则存储数据
                if accuracy_match and ao_mango_match and guiqi_mango_match and jinhuang_mango_match and tainong_mango_match:
                    data['Folder'].append(folder_name)
                    data['Accuracy'].append(float(accuracy_match.group(1)))
                    data['Precision Ao Mango'].append(float(ao_mango_match.group(1)))
                    data['Recall Ao Mango'].append(float(ao_mango_match.group(2)))
                    data['F1 Score Ao Mango'].append(float(ao_mango_match.group(3)))
                    data['Precision GuiQi Mango'].append(float(guiqi_mango_match.group(1)))
                    data['Recall GuiQi Mango'].append(float(guiqi_mango_match.group(2)))
                    data['F1 Score GuiQi Mango'].append(float(guiqi_mango_match.group(3)))
                    data['Precision Jinhuang Mango'].append(float(jinhuang_mango_match.group(1)))
                    data['Recall Jinhuang Mango'].append(float(jinhuang_mango_match.group(2)))
                    data['F1 Score Jinhuang Mango'].append(float(jinhuang_mango_match.group(3)))
                    data['Precision Tainong Mango'].append(float(tainong_mango_match.group(1)))
                    data['Recall Tainong Mango'].append(float(tainong_mango_match.group(2)))
                    data['F1 Score Tainong Mango'].append(float(tainong_mango_match.group(3)))
                    data['Flops'].append(Flops_match.group(1))
                    data['Params'].append(Params_match.group(1))

# 创建一个DataFrame并保存到Excel文件
df = pd.DataFrame(data)
excel_path = 'mango_results.xlsx'
df.to_excel(excel_path, index=False)

# 打开Excel文件并设置字体颜色和列宽
wb = load_workbook(excel_path)
ws = wb.active

# 定义颜色
colors = {
    'black': '00000000',
    'red': 'FFFF0000',
    'blue': 'FF0000FF',
    'green': 'FF008000',
    'purple': 'FF800080'
}

# 设置字体颜色和列宽
column_settings = {
    'A': colors['black'],
    'B': colors['black'],
    'C': colors['red'],
    'D': colors['red'],
    'E': colors['red'],
    'F': colors['blue'],
    'G': colors['blue'],
    'H': colors['blue'],
    'I': colors['green'],
    'J': colors['green'],
    'K': colors['green'],
    'L': colors['purple'],
    'M': colors['purple'],
    'N': colors['purple'],
    'O': colors['black'],
    'P': colors['black'],
}

i=0
for col, color in column_settings.items():
    for cell in ws[col]:
        cell.font = Font(color=color)
    if i== 0:
        ws.column_dimensions[col].width = 15
    else:
        ws.column_dimensions[col].width = 10  # 设置列宽为20，具体宽度可以根据实际情况调整
    i = i+1



# 保存修改后的Excel文件
wb.save(excel_path)

print(f'数据已保存到 {excel_path}，并设置了指定列的字体颜色和列宽')
